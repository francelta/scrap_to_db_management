import sqlite3
from subiendo_web import *
from cocinasplus1 import numbers_to_columns,duplicarExcel,arrowDown,arrowUp,traductorCP,parafrasear,correctorGramatical,getParafraser,numbers_to_columns,get_sku
from openpyxl import load_workbook
import time
import woocommerce
from woocommerce import API
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from utilidades_2 import get_producto
from rytr.languages import Languages
from rytr.tones import Tones
from rytr.usecases import UseCases
from rytr.content import Content
import os
import time
from datetime import datetime
import googletrans
from googletrans import Translator
import pymysql
import random
import urllib
import re
# from wordpress import API as wpAPI

# wpapi = wpAPI(
# url="https://www.pluscookingdeluxe.com/",
# api="wp-json",
# version='wp/v2',
# wp_user="francisco@checkinscan.com",
# wp_pass="Cp999666555*",
# basic_auth = True,
# user_auth = True,
# )
translator = Translator()

toneIds = ["6058207530f7b1000c1c4f86", 
"6058209c30f7b1000c1c4f88", 
"6058223630f7b1000c1c4f96", 
"605820c030f7b1000c1c4f89", 
"60572a639bdd4272b8fe358a", 
"605820d430f7b1000c1c4f8a", 
"605820e330f7b1000c1c4f8b", 
"60572a639bdd4272b8fe358b", 
"60e96f6992161b0013c6ae4a", 
"6058212830f7b1000c1c4f8d", 
"6058213830f7b1000c1c4f8e", 
"6058200830f7b1000c1c4f85", 
"60572a649bdd4272b8fe358c", 
"6058215930f7b1000c1c4f8f", 
"6058216730f7b1000c1c4f90", 
"60ff8d3afc873e000c08e8b2", 
"6064c6679bde74000cea994c", 
"6058219030f7b1000c1c4f92", 
"6058208730f7b1000c1c4f87", 
"605821c030f7b1000c1c4f93", 
"605821cc30f7b1000c1c4f94", 
"605821e030f7b1000c1c4f95",]
wcapi = API(
    url="https://www.pluscookingdeluxe.com/",
    consumer_key="ck_d1fe7be8dc243452980c705caebc57e6fa28fa91",
    consumer_secret="cs_89625e4b5593197f092f91b92da150e97a69359e",
    version="wc/v3",
    timeout=30
)
marcas={'001': 'RIERA', '002': 'BERGHOFF', '003': 'BUGATTI', '004': 'RUFFONI', '005': 'CRISTEL','007': 'RUNBOTT', '008': 'JOSEPH JOSEPH', '009': 'LEKUÈ','010': 'WÜSTHOFF', '011': 'DE BUYER',
        '012': 'BRABANTIA', '013': 'BOSKA', '014': 'KAI','016': 'LEOPOLD VIENA', '017': 'SIMPLE HUMAN', '018':'BUENO MAISON BERGER',
        '019': 'ADE', '020':'BREDEMEIJER', '021': 'BREKA', '022':'KYOCERA', '023':'LE CREUSET',
        '027':'LANGUIOLE', '028':'EMILE HENRY', '029':'MAUVIEL', '030':'PEUGEOT', '031':'COLE & MASON', '032':'GEFU', '033': 'SPIEGELAU', '034':'ADHOC',
        '035': 'EVA SOLO', '036': 'LE JACQUARD FRANCAIS', '037':'MICROPLANE', '038':'SPRING', '039':'JURGEN CHAUMARAT', '040':'EL ARTE DEL OLIVO', '043':'BRA',
        '044':'SMEG', '045':'KITCHEN CRAFT', '046':'ALFI', '048':'TOKYO','050':'BIALETTI','052':'WOLL','053':'NACHTMAN', '054':'VULKANUS','056':'KAMBUKA', '057':'PTM','060':'WILLIAM BOUNDS', '061':'VILLEROY & BOSCH',
        '062':'MIYABI', '063':'EXCELSA', '064':'AEG', '065':'VIN BOUQUET', '066':'MOHA', '067':'MARLUX', '068':'SWELL', '069':'MEMENTO', '070':'TOUCH MEL',
        '071':'ALESSI', '072':'OXO', '073':'PATISSE', '074':'VIVO', '076':'WMF', '077':'BRA', '078':'POMME PIDOU', '079':'GUZZINI', '080':'MR WONDERFUL',
        '081':'KITCHEN AID', '082':'HOMESOLUTIONS', '083':'LA ROCHERE', '084':'ROSTI', '085':'OIL +', '086':'KUHN RIKON', '087':'REISENTHEL', '088':'PEBBLY', '089':'KUCHENPROFI',
        '090':'PODEVACHE', '091':'CILIO', '093':'MARINA BUSSINESS', '094':'ESTEBAN', '096':'PIP STUDIO', '097':'STAUB'}


marcas_en_woo={"Ade":999,
"Adhoc":1011,
"Aeg":1032,
"Alessi":1039,
"Alfi":1020,
"le_jaqguard":988,
"Bialetti":1022,
"Boska":994,
"Bra":1018,
"Brabantia":993,
"Bredemeijer":1000,
"Breka":1001,
"Bueno maison berger":998,
"Bugatti":989,
"Cilio":1057,
"Cole &amp; mason":1008,
"Cristel":348,
"De buyer":992,
"El arte del olivo":1017,
"Emile henry":1005,
"Esteban":1059,
"Eva solo":1012,
"Excelsa":1031,
"Gefu":1009,
"Guzzini":1045,
"Homesolutions":1048,
"Joseph Joseph":143,
"Jurgen chaumarat":1016,
"Kai":995,
"Kambuka":1026,
"Kitchen aid":1047,
"Kitchen craft":1019,
"Kuchenprofi":1055,
"Kuhn rikon":1052,
"Kyocera":1002,
"La rochere":1049,
"Languiole":1004,
"Le creuset":1003,
"Le jacquard francais":1013,
"Lekue":314,
"Leopold viena":996,
"Marina bussiness":1058,
"Marlux":1035,
"Mauviel":1006,
"Memento":1037,
"Microplane":1014,
"Miyabi":1030,
"Moha":1034,
"Mr wonderful":1046,
"Nachtman":1024,
"Oil +":1051,
"Oxo":1040,
"Patisse":1041,
"Pebbly":1054,
"Peugeot":1007,
"Pip studio":1060,
"Podevache":1056,
"Pomme pidou":1044,
"Ptm":1027,
"Reisenthel":1053,
"Riera":987,
"Rosti":1050,
"Ruffoni":349,
"Runbott":990,
"Simple human":997,
"Smeg":315,
"Spiegelau":1010,
"Spring":1015,
"Staub":1061,
"Swell":1036,
"Tokyo":1021,
"Touch mel":1038,
"Villeroy &amp; bosch":1029,
"Vin bouquet":1033,
"Vivo":1042,
"Vulkanus":1025,
"William bounds":1028,
"Wmf":1043,
"Woll":1023,
"Wüsthoff":991}




categorias={'001':'Cocinar>Baterias y set',
'002':'Cocinar>Ollas',
'003':'Cocinar>Cacerolas',
'004':'Cocinar>Cacillos',
'005':'Cocinar>Del chef',
'006':'Cocinar>Ollas a presión',
'007':'Cocinar>Sartenes y paelleras',
'008':'Cocinar>Parrillas y planchas',
'009':'Cocinar>Para microondas',
'010':'Cocinar>Tapaderas',
'011':'Cocinar>Mangos y asas',

'012':'Cocinar>Textil>Manoplas',
'013':'Cocinar>Textil>Paños',
'014':'Cocinar>Textil>Delantales',

'015':'Cocinar>Cocina especial>Vaporeras',
'016':'Cocinar>Cocina especial>Fondue',
'017':'Cocinar>Cocina especial>Wok',
'018':'Cocinar>Cocina especial>Raclettes',
'019':'Cocinar>Cocina especial>Rustidera y asados',
'020':'Cocinas>Cocina especial>Miniaturas',
'021':'Cocinar>Hornear>Moldes',
'022':'Cocinar>Hornear>Fuentes y bandejas',
'023':'Cocinar>Hornear>Decoración',
'024':'Cocinar>Hornear>Accesorios',
'025':'Cocinar>Hornear>Panadería y respostería>Utensilios',
'026':'Cocinar>Hornear>Panadería y respostería>Cortadores y rodillos',
'027':'Cocinar>Hornear>Panadería y respostería>Decoración',
'028':'Cortar>Cuchillos>Santoku',
'029':'Cortar>Cuchillos>Chef',
'030':'Cortar>Cuchillos>Japoneses',
'031':'Cortar>Cuchillos>Carne',
'032':'Cortar>Cuchillos>Pescados y mariscos',
'033':'Cortar>Cuchillos>Fruta y vegetal',
'034':'Cortar>Cuchillos>Pan y queso',
'035':'Cortar>Set de cuchillos',
'036':'Cortar>Bloques de cuchillos',
'037':'Cortar>Afiladores y chairas',
'038':'Cortar>Tablas de cortar',
'039':'Cortar>Mandolinas',
'040':'Cortar>Tijeras',
'041':'Cortar>Peladores',
'042':'Cortar>Ralladores',


'043':'Electrodomésticos>Cafeteras eléctricas',
'044':'Electrodomésticos>Hervidores',
'045':'Electrodomésticos>Batidoras y licuadoras',
'046':'Electrodomésticos>Exprimidores',
'047':'Electrodomésticos>Tostadores',
'048':'Electrodomésticos>Básculas',
'049':'Electrodomésticos>Procesadores de alimentos',
'050':'Electrodomésticos>Deshidratadores',
'051':'Electrodomésticos>Raclettes',
'052':'Electrodomésticos>Plachas',


'053':'Mesa y Bar>Vajilla>Set',
'054':'Mesa y Bar>Vajilla>Platos',
'055':'Mesa y Bar>Vajilla>Boles',
'056':'Mesa y Bar>Vajilla>Para servir',
'057':'Mesa y Bar>Vajilla>Especiales',

'058':'Mesa y Bar>Cristalería>Vasos',
'059':'Mesa y Bar>Cristalería>Vino Tinto',
'060':'Mesa y Bar>Cristalería>Vino Blanco',
'061':'Mesa y Bar>Cristalería>Cava',
'062':'Mesa y Bar>Cristalería>Decantadores',
'063':'Mesa y Bar>Cristalería>Licores y otros',
'064':'Mesa y Bar>Cristalería>Jarras',

'065':'Mesa y Bar>Cuberterías>Sets',
'066':'Mesa y Bar>Cuberterías>Para servir',
'067':'Mesa y Bar>Cubertterías>Especiales',

'068':'Mesa y Bar>Café y té>Sets',
'069':'Mesa y Bar>Café y té>Tazas',
'070':'Mesa y Bar>Café y té>Cafeteras',
'071':'Mesa y Bar>Café y té>Teteras',
'072':'Mesa y Bar>Café y té>Azucareros y cremeras',
'073':'Mesa y Bar>Café y té>Accesorios',

'074':'Mesa y Bar>Textil>Manteles',
'075':'Mesa y Bar>Textil>Individuales',
'076':'Mesa y Bar>Textil>Servilletas',
'077':'Mesa y Bar>Textil>Caminos de mesa',
'078':'Mesa y Bar>Accesorios de mesa>Salvamanteles',
'079':'Mesa y Bar>Accesorios de mesa>Servilleteros',
'080':'Mesa y Bar>Accesorios de mesa>Candelabros',

'081':'Mesa y Bar>Set y condimentos>Sal y pimienta',
'082':'Mesa y Bar>Set y condimentos>Aceite y vinagre',
'083':'Mesa y Bar>Set y condimentos>Hierbas y especias',
'084':'Mesa y Bar>Set y condimentos>Queso',

'085':'Mesa y Bar>Termos y botellas',

'086':'Mesa y Bar>Bar>Coctelería',
'087':'Mesa y Bar>Bar>Moldes de hielo',
'088':'Mesa y Bar>Bar>Enfriadores y cubiteras',
'089':'Mesa y Bar>Bar>Accesorios',

'090':'Mesa y Bar>Muebles auxiliares',

'091':'Exterior>Cocinar',
'092':'Exterior>Utensilios barbacoa',
'093':'Exterior>Termos',
'094':'Exterior>Picnic',
'095':'Exterior>Vajillas',
'096':'Exterior>Cuberterías',
'097':'Exterior>Copas y vasos',

'098':'Utensilios>Esenciales>Sets',
'099':'Utensilios>Esenciales>Cucharas y cucharones',
'100':'Utensilios>Esenciales>Espátulas y pinceles',
'101':'Utensilios>Esenciales>Pinzas',
'102':'Utensilios>Esenciales>Batidores',
'103':'Utensilios>Esenciales>Coladores y escurridores',
'104':'Utensilios>Esenciales>Guantes y protectores',
'105':'Utensilios>Esenciales>Ralladores',
'106':'Utensilios>Esenciales>Peladores',
'107':'Utensilios>Esenciales>Abridores',

'108':'Utensilios>Especializados>Cítricos',
'109':'Utensilios>Especializados>Huevo',
'110':'Utensilios>Especializados>Morteros y molinillos',
'111':'Utensilios>Especializados>Carnes, marsicos y pescados',
'112':'Utensilios>Especializados>Frutas y verduras',
'113':'Utensilios>Especializados>Helados',
'114':'Utensilios>Especializados>Pizza y pasta',
'115':'Utensilios>Especializados>Quesos',
'116':'Utensilios>Especializados>Sopletes y sifones',

'117':'Utensilios>Organizadores y medidores>Medidores',
'118':'Utensilios>Organizadores y medidores>Boles',
'119':'Utensilios>Organizadores y medidores>Temporizadores',
'120':'Utensilios>Organizadores y medidores>Termómetros',
'121':'Utensilios>Organizadores y medidores>Básculas',
'122':'Utensilios>Organizadores y medidores>Almacenamiento',

'123':'Utensilios>Limpieza>Cepillos',
'124':'Utensilios>Limpieza>Dispensadores',
'125':'Utensilios>Limpieza>Accesorios',

'126':'Utensilios>Cubos de basura',
'127':'Utensilios>Otros>Baño>Toalla',
'128':'Utensilios>Otros>Baño>Albornoz',
'129':'Utensilios>Otros>Baño>Otros',
}



categorias_={'145':'>Planchas',
'146':'>Raclettes',
'147':'>Deshidratadores',
'148':'>Procesadores de alimentos',
'149':'>Básculas',
'155':'>Exterior',
'156':'>Copas y vasos',
'157':'>Cuberterías',
'158':'>Vajillas',
'159':'>Picnic',
'160':'>Termos',
'161':'>Utensilios barbacoa',
'162':'>Cocinar',
'163':'>Mesa y Bar',
'164':'>Muebles auxiliares',
'165':'>Bar',
'166':'>Accesorios',
'167':'>Enfriadores y cubiteras',
'168':'>Moldes de hielo',
'169':'>Coctelería',
'170':'>Termos y botellas',
'171':'>Set y condimentos',
'172':'>Queso',
'173':'>Hierbas y especias',
'174':'>Aceite y vinagre',
'175':'>Sal y pimienta',
'176':'>Accesorios de mesa',
'177':'>Textil',
'178':'>Candelabros',
'179':'>Servilleteros',
'180':'>Salvamanteles',
'181':'>Caminos de mesa',
'182':'>Servilletas',
'183':'>Individuales',
'184':'>Manteles',
'185':'>Café y té',
'186':'>Accesorios',
'187':'>Azucareros y cremeras',
'188':'>Teteras',
'189':'>Cafeteras',
'190':'>Tazas',
'191':'>Sets',
'192':'>Cuberterías',
'193':'>Especiales',
'194':'>Para servir',
'195':'>Sets',
'196':'>Cristalería',
'197':'>Jarras',
'198':'>Licores y otros',
'199':'>Decantadores',
'200':'>Cava',
'201':'>Vino blanco',
'202':'>Vino tinto',
'203':'>Vasos',
'204':'>Vajilla',
'205':'>Especiales',
'206':'>Para servir',
'207':'>Boles',
'208':'>Platos',
'209':'>Set',
'210':'>Utensilios',
'211':'>Cubos de basura',
'212':'>Limpieza',
'213':'>Accesorios',
'214':'>Dispensadores',
'215':'>Cepillos',
'216':'>Organizadores y medidores',
'218':'>Almacenamiento',
'219':'>Básculas',
'220':'>Termómetros',
'221':'>Temporizadores',
'222':'>Boles',
'223':'>Medidores',
'224':'>Especializados',
'225':'>Sopletes y sifones',
'226':'>Quesos',
'227':'>Pizza y pasta',
'228':'>Helados',
'229':'>Frutas y verduras',
'230':'>Carnes, mariscos y pescados',
'231':'>Morteros y molinillos',
'232':'>Huevo',
'233':'>Cítricos',
'234':'>Esenciales',
'235':'>Abridores',
'236':'>Peladores',
'237':'>Ralladores',
'238':'>Guantes y protectores',
'239':'>Coladores y escurridores',
'240':'>Batidores',
'241':'>Pinzas',
'242':'>Espátulas y pinceles',
'243':'>Cucharas y cucharones',
'244':'>Sets',
'245':'>Cortar',
'246':'>Ralladores',
'247':'>Peladores',
'248':'>Tijeras',
'249':'>Mandolinas',
'250':'>Tablas de cortar',
'251':'>Afiladores y chairas',
'252':'>Bloques de cuchillos',
'253':'>Set de cuchillos',
'254':'>Cuchillos',
'255':'>Pan y queso',
'256':'>Fruta y vegetal',
'257':'>Pescados y mariscos',
'258':'>Carne',
'259':'>Japoneses',
'260':'>Chef',
'261':'>Santoku',
'262':'>Hornear',
'263':'>Panadería y repostería',
'264':'>Cortadores y rodillos',
'265':'>Decoración',
'266':'>Utensilios',
'267':'>Accesorios',
'268':'>Decoración',
'269':'>Fuentes y bandejas',
'270':'>Moldes',
'271':'>Cocinar',
'272':'>Cocina especial',
'273':'>Miniaturas',
'274':'>Rustidera y asados',
'275':'>Raclette',
'276':'>Wok',
'277':'>Fondue',
'278':'>Vaporeras',
'279':'>Textil',
'280':'>Delantales',
'281':'>Paños',
'282':'>Manoplas',
'283':'>Mangos y asas',
'284':'>Tapaderas',
'285':'>Para microondas',
'286':'>Parrillas y planchas',
'287':'>Sartenes y paelleras',
'288':'>Ollas a presión',
'289':'>Del chef',
'290':'>Cacillos',
'291':'>Cacerolas',
'292':'>Ollas',
'293':'>Baterias y set',
'316':'>Electrodomesticos',
'317':'>Tostadores',
'318':'>Hervidores',
'319':'>Exprimidores',
'320':'>Batidoras y licuadoras',
'321':'>Cafeteras eléctricas',
'601':'>Cooking',
'602':'>Set',
'603':'>Pans',
'604':'>saucepans',
'605':'>Special cooking',
'606':'>Fondue',
'607':'>Miniatures',
'608':'>Raclette',
'609':'>Roasters',
'610':'>Steamers',
'611':'>Wok',
'612':'>Chef´s',
'613':'>Handles',
'614':'>Pots',
'615':'>Pressure cookers',
'616':'>Microwaveable',
'617':'>Grills and plancha',
'618':'>Pans',
'619':'>lids',
'620':'>Textil',
'621':'>Aprons',
'622':'>Gloves',
'623':'>Kitchen towels',
'624':'>Cutting',
'625':'>Sharpeners',
'626':'>Knife blocks',
'627':'>Knifes',
'628':'>Meat',
'629':'>Chef',
'630':'>Fruit and vegetable',
'631':'>Japanese knives',
'632':'>Bread and cheese',
'633':'>Fish and shellfish',
'634':'>Santoku',
'635':'>Mandoline',
'636':'>Peelers',
'637':'>Graters',
'638':'>Knife set',
'639':'>Chopping boards',
'640':'>Scissors',
'641':'>Home appliances',
'642':'>Mixers and blenders',
'643':'>Scales',
'644':'>Electric coffee makers',
'645':'>Dehydrators',
'646':'>Kettles',
'647':'>Planchas',
'648':'>Food processors',
'649':'>Raclettes',
'650':'>Toaster',
'652':'>Exterior',
'653':'>Cooking exterior',
'654':'>Cups and glasses',
'655':'>Exterior Cutlery',
'656':'>Picnic',
'657':'>Thermos',
'658':'>Barbecue utensils',
'659':'>Tableware',
'660':'>Baking',
'661':'>Accessories',
'662':'>Decor',
'663':'>Molds',
'664':'>Trays',
'665':'>Bakery',
'666':'>Cutters and rollers',
'667':'>Decor',
'668':'>Utensils',
'669':'>Bar',
'670':'>Table accessories',
'671':'>Chandeliers',
'672':'>Mat',
'673':'>Napkin rings',
'674':'>Bar',
'675':'>Accessories',
'676':'>Cocktails',
'677':'>Coolers and ice buckets',
'678':'>Ice molds',
'679':'>Coffee and tea',
'680':'>Accessories',
'681':'>Sugar bowls and creamers',
'682':'>Coffee makers',
'683':'>Sets',
'684':'>Cups',
'685':'>Teapots',
'686':'>Glassware',
'687':'>Cava',
'688':'>Decanters',
'689':'>Jugs',
'690':'>Liquors and others',
'691':'>Glasses',
'692':'>White wine',
'693':'>Red wine',
'694':'>Cutlery',
'695':'>Specials',
'696':'>Serving',
'697':'>Sets',
'698':'>Auxiliary furniture',
'699':'>Condiments and set',
'700':'>Oil and vinegar',
'701':'>Herbs and spices',
'702':'>Cheese',
'703':'>Salt and pepper',
'704':'>Thermos and bottles',
'705':'>Table Textil',
'706':'>Table runners',
'707':'>Individual',
'708':'>Tablecloths',
'709':'>Napkins',
'710':'>Crockery',
'711':'>Bowls',
'712':'>Specials',
'713':'>Serving',
'714':'>Dishes',
'715':'>Set',
'716':'>Utensils',
'717':'>Bin',
'718':'>Essentials',
'719':'>Openers',
'720':'>Whisk',
'721':'>Strainers and drainers',
'722':'>Spoons and ladles',
'723':'>Spatulas and brushes',
'724':'>Gloves and protectors',
'725':'>Peelers',
'726':'>Tongs',
'727':'>Graters',
'728':'>Sets',
'729':'>Specialized',
'730':'>Meat, seafood and fish',
'731':'>Citrus',
'732':'>Fruits and vegetables',
'733':'>Ice Creams',
'734':'>Egg',
'735':'>Mortars and grinders',
'736':'>Pizza and pasta',
'737':'>Cheeses',
'738':'>Blowtorch',
'739':'>Cleaning',
'740':'>Cleaning accesories',
'741':'>Brushes',
'742':'>Dispensers',
'743':'>Kitchen organizers',
'744':'>Storage',
'745':'>Scales',
'746':'>Bowls',
'747':'>Measurer',
'748':'>Timers',
'749':'>Thermometers',
'750':'>Juicers',
'884':'>Cocinas',
'885':'>Cocina especial',
'886':'>Miniaturas',
'887':'>Panadería y respostería',
'888':'>Utensilios',
'925':'>Cocinar>Cacerolas',
'926':'>Cocinar>Ollas',
'927':'>Cocinar>Sartenes y paelleras',
'928':'>Hornear>Fuentes y bandejas',
'929':'>Cocinar>Cocina especial>Rustidera y asados',
'930':'>Cocinar>Cacillos',
'931':'>Mesa y Bar>Vajilla>Especiales',
'932':'>Cocinar>Cocina especial>Fondue',
'933':'>Cocinas>Cocina especial>Miniaturas',
'934':'>Cocinar>Cocina especial>Wok',
'935':'>Cocinar>Del chef',
'936':'>Mesa y Bar>Accesorios de mesa>Candelabros',
'937':'>Mesa y Bar>Accesorios de mesa>Salvamanteles',
'938':'>Utensilios>Esenciales>Coladores y escurridores',
'939':'>Cocinar>Baterias y set',
'940':'>Mesa y Bar>Vajilla>Para servir',
'941':'>Cocinar>Tapaderas',
'942':'>Mesa y Bar>Bar>Enfriadores y cubiteras',
'943':'>Utensilios>Esenciales>Sets',
'944':'>Hornear>Panadería y respostería>Utensilios',
'945':'>Utensilios>Esenciales>Ralladores',
'946':'>Utensilios>Organizadores y medidores>Almacenamiento',
'947':'>Hornear>Moldes',
'948':'>Cacerolas',
'949':'>Ollas',
'950':'>Sartenes y paelleras',
'951':'>Fuentes y bandejas',
'952':'>Cocina especial',
'953':'>Rustidera y asados',
'954':'>Cacillos',
'955':'>Vajilla',
'956':'>Especiales',
'957':'>Fondue',
'958':'>Cocina especial',
'959':'>Miniaturas',
'960':'>Wok',
'961':'>Del chef',
'962':'>Accesorios de mesa',
'963':'>Candelabros',
'964':'>Salvamanteles',
'965':'>Esenciales',
'966':'>Coladores y escurridores',
'967':'>Baterias y set',
'968':'>Para servir',
'969':'>Tapaderas',
'970':'>Enfriadores y cubiteras',
'971':'>Sets',
'972':'>Panadería y respostería',
'973':'>Utensilios',
'974':'>Ralladores',
'975':'>Organizadores y medidores',
'976':'>Almacenamiento',
'977':'>Moldes',
'983':'>Miniaturas',
'984':'>Cocinas',
'985':'>Cocina especial',
'1137':'>Otros',
'1138':'>Others',
'1139':'>Baño',
'1140':'>Toallas',
'1141':'>Albornoces',
'1142':'>Otros',
'1143':'>Bathroom',
'1144':'>Bathrobes',
'1145':'>Others',}

'''
    se crean dos bases de datos, una en castellano y otra en inglés, 
   para traducir y parafrasear se usa rytr, scrapeamos tanto las fotos como las descripciones en bruto (a traducir y parafrasear)
   
   usando el wp-cli se crean las categorias y se suben las fotos, se crea un post por cada producto y se suben los campos de la base de datos
   
'''

def create_marca_smeg(marca):
    conn = sqlite3.connect('cocinasplus.db')
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS "+marca+" (sku text, ean text, marca text, nombre text, nombre_foto text, longitud text, anchura text, altura text, peso text, precio_coste real, precio_venta real, descripcion_bruto text, descripcion text, stock text, categoria text)")
    conn.commit()
    conn.close()

def create_trade_mark_smeg(marca):
    conn = sqlite3.connect('cocinasplus_en.db')
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS "+marca+" (sku text, ean text, marca text, nombre text, nombre_foto text, longitud text, anchura text, altura text, peso text, precio_coste real, precio_venta real, descripcion_bruto text, descripcion text, stock text, categoria text)")
    conn.commit()
    conn.close()



def insert_campos(marca,sku,ean,marca1,nombre,nombre_foto,longitud,anchura,altura,peso,precio_coste,precio_venta,descripcion_bruto,descripcion,stock,categoria):
    conn = sqlite3.connect('cocinasplus.db')
    c = conn.cursor()
    
    c.execute("INSERT INTO "+marca+" VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(sku,ean,marca1,nombre,nombre_foto,longitud,anchura,altura,peso,precio_coste,precio_venta,descripcion_bruto,descripcion,stock,categoria))
    conn.commit()
    conn.close()
    
def insert_campos_en(marca,sku,ean,marca1,nombre,nombre_foto,longitud,anchura,altura,peso,precio_coste,precio_venta,descripcion_bruto,descripcion,stock,categoria):
    conn = sqlite3.connect('cocinasplus_en.db')
    c = conn.cursor()
    
    c.execute("INSERT INTO "+marca+" VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(sku,ean,marca1,nombre,nombre_foto,longitud,anchura,altura,peso,precio_coste,precio_venta,descripcion_bruto,descripcion,stock,categoria))
    conn.commit()
    conn.close()

def update_campo(marca,sku,campo,valor):
    conn = sqlite3.connect('cocinasplus.db')
    c = conn.cursor()
    c.execute("UPDATE "+marca+" SET "+campo+" = '"+valor+"' WHERE sku = '"+sku+"'")
    conn.commit()
    conn.close()
    
def update_campo_en(marca,sku,campo,valor):
    conn = sqlite3.connect('cocinasplus_en.db')
    c = conn.cursor()
    c.execute("UPDATE "+marca+" SET "+campo+" = '"+valor+"' WHERE sku = '"+sku+"'")
    conn.commit()
    conn.close()

def query_campo(marca,campo,valor):
    conn = sqlite3.connect('cocinasplus.db')
    c = conn.cursor()
    c.execute("SELECT * FROM "+marca+" WHERE "+campo+" = '"+valor+"'")
    conn.commit()
    conn.close()
    return c.fetchall()

def query_campo_en(marca,campo,valor):
    conn = sqlite3.connect('cocinasplus_en.db')
    c = conn.cursor()
    c.execute("SELECT * FROM "+marca+" WHERE "+campo+" = '"+valor+"'")
    conn.commit()
    conn.close()
    return c.fetchall()


def check_marca_simple(sku,excel_marca_simple):
    
    wb=load_workbook(excel_marca_simple)
    ws=wb.active
    
    a=ws['A2':'A'+str(ws.max_row)]
    
    for i in range(len(a)-1):
        if sku == a[i][0].value:
            nombre_producto=str(ws['C'+str(i+2)].value)
            
            ean=str(ws['E'+str(i+2)].value)
            longitud=str(ws['M'+ str(i+2)].value)
            altura=str(ws['O'+str(i+2)].value)
            anchura=str(ws['L'+str(i+2)].value)
            peso=str(ws['J'+str(i+2)].value)
            descripcion_bruto=str(ws['R'+str(i+2)].value)
            descripcion_bruto_en=str(ws['P'+str(i+2)].value)
            categoria=str(ws['D'+str(i+2)].value)
            precio_coste=round(float(ws['G'+str(i+2)].value),1)
            precio_venta=round(float(ws['H'+str(i+2)].value),1)
            configurable=False
            break
        else:
            nombre_producto=''
            longitud=''
            altura=''
            anchura=''
            peso=''
            descripcion_bruto=''
            descripcion_bruto_en=''
            categoria=''
            configurable=True
    
            
    return (configurable,nombre_producto,ean,longitud,altura,anchura,peso,descripcion_bruto,descripcion_bruto_en,categoria,precio_coste,precio_venta)
            
            
def parafrasear_nombre(nombre):
    
    usecase = UseCases.get(id='60583ac98c0a4a000c69c96f')
    usecase = usecase["data"]

    content = Content.generate(
        user_id=1,
        language_id="607adad66f8fe5000c1e636e",
        tone_id= "6058212830f7b1000c1c4f8d",
        # "6058207530f7b1000c1c4f86",Appreciative
        # '6058213830f7b1000c1c4f8e','Enthusiastic'
        usecase_id=usecase["_id"],
        input_contexts={
            usecase["contextInputs"][0]["keyLabel"]: nombre
        },
    )
    print(content)
    return(content['data'][0]['text'])     

def parafrasear_descripcion(descripcion_bruto,toneId):
    
    if len(descripcion_bruto) > 250:
        descripcion_bruto_1=descripcion_bruto[:249]
        descripcion_bruto_2=descripcion_bruto[249:498]
        try:
            usecase = UseCases.get(id='60928476a9c7620013304e89')
            usecase = usecase["data"]
            
            content = Content.generate(
                user_id=1,
                language_id="607adad66f8fe5000c1e636e",
                tone_id=toneId,
                usecase_id=usecase["_id"],
                input_contexts={
                    usecase["contextInputs"][0]["keyLabel"]: descripcion_bruto_1
                },
            )
            
            try:
                descripcion_1=content['data'][0]['text']
            except:
                descripcion_1=content['data']['text']
            time.sleep(0.5)
        except:
            descripcion_1=descripcion_bruto_1
        try:
            descripcion_en_1=translator.translate(descripcion_1,src='es',dest='en')
            descripcion_en_1=descripcion_en_1.text
        except:
            descripcion_en_1=descripcion_1
        time.sleep(0.1)
            
        try:
            usecase = UseCases.get(id='60928476a9c7620013304e89')
            usecase = usecase["data"]
            
            content = Content.generate(
                user_id=1,
                language_id="607adad66f8fe5000c1e636e",
                tone_id="60ff8d3afc873e000c08e8b2",
                usecase_id=usecase["_id"],
                input_contexts={
                    usecase["contextInputs"][0]["keyLabel"]: descripcion_bruto_2
                },
            )
            try:
                descripcion_2=content['data'][0]['text']
            except:
                descripcion_2=content['data']['text']
            time.sleep(0.5)
        except:
            descripcion_2=descripcion_bruto_2
        try:
            descripcion_en_2=translator.translate(descripcion_2,src='es',dest='en')
            descripcion_en_2=descripcion_en_2.text
        except:
            descripcion_en_2=descripcion_2
        descripcion=descripcion_1+' '+descripcion_2
        descripcion_en=descripcion_en_1+' '+descripcion_en_2
    else:
        time.sleep(0.1)
        try:
            usecase = UseCases.get(id='60928476a9c7620013304e89')
            print('1')
            usecase = usecase["data"]
            print('2')
            time.sleep(0.1)
            content = Content.generate(
                user_id=1,
                language_id="607adad66f8fe5000c1e636e",
                tone_id="60ff8d3afc873e000c08e8b2",
                usecase_id=usecase["_id"],
                input_contexts={
                    usecase["contextInputs"][0]["keyLabel"]: descripcion_bruto
                },
            )
            print('3')
            try:
                descripcion=content['data'][0]['text']
            except:
                descripcion=content['data']['text']
            time.sleep(0.5)
        except:
            descripcion=descripcion_bruto
            
    return(descripcion)
 
def find_longest_word(sentence):
    words = sentence.split()
    try:
        longest_word = max(words, key=len)
        sentence=sentence.replace(longest_word, '')
    except:
        longest_word = sentence
        
    
    return (longest_word,sentence) 

def find_cat(word):
    for i in categorias_: 
        palabra=word.upper()
        linea=categorias_[i].split('>')[-1].upper()
        if linea.find(palabra.upper()) != -1:
            return(i)
            
        else:
            
            palabra=word[:len(word)-1]
            if linea.find(palabra.upper()) != -1:
                return(i)
            else:
                
                palabra=word[:len(word)-2]
                if linea.find(palabra.upper()) != -1:
                    return(i)

def parafrasear_descripcion_corta(descripcion_bruto,toneId):
    
    if len(descripcion_bruto) > 250:
        descripcion_bruto_1=descripcion_bruto[:249]
        descripcion_bruto_2=descripcion_bruto[249:498]
        try:
            usecase = UseCases.get(id='60928752a9c7620013304ea1')
            usecase = usecase["data"]
            
            content = Content.generate(
                user_id=1,
                language_id="607adad66f8fe5000c1e636e",
                tone_id=toneId,
                usecase_id=usecase["_id"],
                input_contexts={
                    usecase["contextInputs"][0]["keyLabel"]: descripcion_bruto_1
                },
            )
            
            try:
                descripcion_1=content['data'][0]['text']
            except:
                descripcion_1=content['data']['text']
            time.sleep(0.5)
        except:
            descripcion_1=descripcion_bruto_1
        try:
            descripcion_en_1=translator.translate(descripcion_1,src='es',dest='en')
            descripcion_en_1=descripcion_en_1.text
        except:
            descripcion_en_1=descripcion_1
        time.sleep(0.1)
            
        try:
            usecase = UseCases.get(id='60928752a9c7620013304ea1')
            usecase = usecase["data"]
            
            content = Content.generate(
                user_id=1,
                language_id="607adad66f8fe5000c1e636e",
                tone_id="60ff8d3afc873e000c08e8b2",
                usecase_id=usecase["_id"],
                input_contexts={
                    usecase["contextInputs"][0]["keyLabel"]: descripcion_bruto_2
                },
            )
            try:
                descripcion_2=content['data'][0]['text']
            except:
                descripcion_2=content['data']['text']
            time.sleep(0.5)
        except:
            descripcion_2=descripcion_bruto_2
        try:
            descripcion_en_2=translator.translate(descripcion_2,src='es',dest='en')
            descripcion_en_2=descripcion_en_2.text
        except:
            descripcion_en_2=descripcion_2
        descripcion=descripcion_1+' '+descripcion_2
        descripcion_en=descripcion_en_1+' '+descripcion_en_2
    else:
        time.sleep(0.1)
        try:
            usecase = UseCases.get(id='60928752a9c7620013304ea1')
            print('1')
            usecase = usecase["data"]
            print('2')
            time.sleep(0.1)
            content = Content.generate(
                user_id=1,
                language_id="607adad66f8fe5000c1e636e",
                tone_id="60ff8d3afc873e000c08e8b2",
                usecase_id=usecase["_id"],
                input_contexts={
                    usecase["contextInputs"][0]["keyLabel"]: descripcion_bruto
                },
            )
            print('3')
            try:
                descripcion=content['data'][0]['text']
            except:
                descripcion=content['data']['text']
            time.sleep(0.5)
        except:
            descripcion=descripcion_bruto
            
    return(descripcion)
          
def get_from_smeg(sku):
    
    excel='smeg/smeg.xlsx'
    wb = load_workbook(excel)
    ws = wb.active
    print("conseguir el nombre")
    a_=ws['A6':'A'+str(ws.max_row)] #sku1
    b_=ws['B6':'B'+str(ws.max_row)] #sku2
    c_=ws['C6':'C'+str(ws.max_row)] #sku3
    f_=ws['F6':'F'+str(ws.max_row)] #nombre
    
    
    g_=ws['G6':'G'+str(ws.max_row)] #precio_costo
    
    nombre=''
    precio_costo=''
    largo=0
    anchos=0
    
    for i in range(0,int(ws.max_row)-1):
       
        sku_smeg=['044.'+str(a_[i][0].value), '044.'+str(b_[i][0].value), '044.'+str(c_[i][0].value)]
       
        
            
        if sku==sku_smeg[0] or sku==sku_smeg[1] or sku==sku_smeg[2] or sku in sku_smeg[0] or sku in sku_smeg[1] or sku in sku_smeg[2]:
            
            crudo=str(f_[i][0].value)
            nombre=crudo
            crudo=crudo.split('(')[0]
            print(crudo)
            if 'x' in crudo:
                ancho=float(crudo.split('x')[-1])/100
                crudo=crudo.split('x')[0]
               
                largo=float(crudo.split(' ')[-2])/100
                
            else:
                ancho=0
                largo=float(crudo.split(' ')[-1])/100
                
            print(f'nombre: {nombre}')
            print(f'largo: {largo}')
            print(f'ancho: {ancho}')
            precio_costo=float(g_[i][0].value)*1.21
            
            precio_costo=str(round(precio_costo,2))
            print(f'precio costo: {precio_costo}')
            
            wb.close()
            
            break
     
    print("todo bien") 
    return(nombre, precio_costo,largo,ancho)

    
def get_from_smeg_web(sku):
    
   
    path = '/Users/Fran/Desktop/cocinasplus/chromedriver'
    website='https://www.smeg.es/producto/'
   
    options = Options()
   
    options.add_argument("start-maximized")
   
    options.add_argument('headless') 

    driver = webdriver.Chrome(path)
   
    
    print("abro el navegador...")
    
    while True:
        try:
            driver.get(website+str(sku[4:].upper()))
            
            break
        except:
            print("no se pudo cargar la pagina")
            time.sleep(1)
    time.sleep(0.5)
    
    nombre_1=driver.find_element(by=By.XPATH, value='/html/body/div[1]/div[5]/div[1]/div[1]/h1/span[1]').text
    try:
        nombre_2=driver.find_element(by=By.XPATH, value='/html/body/div[1]/div[5]/div[1]/div[1]/h1/span[2]').text
    except:
        nombre_2=''
    nombre=nombre_1+' '+nombre_2
    print(nombre)
    
    descripcion=driver.find_element(by=By.XPATH, value='/html/body/div[1]/div[6]/div[2]/div[1]/section/div/div/div[2]/div/div').text
    print(descripcion)
    
    imagen=driver.find_element(by=By.XPATH, value='/html/body/div[1]/div[5]/div[1]/div[3]/div[2]/div/div/div[2]/div[1]/img').get_attribute('src')
    
    print(imagen)
    pathImage='/Users/Fran/Desktop/cocinasplus/smeg/fotos/'+sku+'.jpg'
    
    print(pathImage)
    try:
        opener=urllib.request.build_opener()
        opener.addheaders=[('User-Agent','Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1941.0 Safari/537.36')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(imagen, pathImage)
        print("foto descargada")
    
    except Exception as e:
        print(e)
            
    driver.close()
        
    
    try:
        toneId=random.choice(toneIds)
        descripcion=parafrasear_descripcion(descripcion, toneId)
        print(descripcion)
           
    except:
        try:
            time.sleep(0.5)
            toneId=random.choice(toneIds)
            descripcion=parafrasear_descripcion(descripcion, toneId)
            print(descripcion)
        except:
            descripcion=descripcion
            pass
    try:
        toneId=random.choice(toneIds)
        contador=0
        descripcion_corta=parafrasear_descripcion_corta(descripcion, toneId)
        print(len(descripcion_corta))
        while len(descripcion_corta)>300 and contador < 5:
            toneId=random.choice(toneIds)
            descripcion_corta=parafrasear_descripcion_corta(descripcion, toneId)
            print(len(descripcion_corta))
            contador=contador+1
        
        print(descripcion_corta)
    except:
        descripcion_corta=''
        pass
    
    
    
    return(nombre,descripcion, descripcion_corta)
    
def corregir_texto(texto):
    
    path = '/Users/Fran/Desktop/cocinasplus/chromedriver'
    website='https://www.correctoronline.es/'
    options = Options()
    options.add_argument("start-maximized")
    options.add_argument('headless')
    driver = webdriver.Chrome(path)
    print("abro el navegador...")
    while True:
        try:
            driver.get(website)
            break
        except:
            print("no se pudo cargar la pagina")
            time.sleep(1)
    time.sleep(2)
    pon_texto='/html/body/div/div[3]/div[2]'
    busqueda=driver.find_element(by=By.XPATH, value=pon_texto).click()
    texto=driver.find_element(by=By.XPATH, value='/html/body/div[2]/div[3]/div[2]').send_keys(texto)
    boton='/html/body/div[2]/div[3]/form/button'
    pulsar=driver.find_element(by=By.XPATH, value=boton).click()
    texto_corregido=driver.find_element(by=By.XPATH, value='/html/body/div[2]/div[3]/div/div[1]').text
    texto_corregido=str(texto_corregido)
    driver.close()
    return(texto_corregido)
    
def build_marca_smeg(excel, nombre_marca, web,buscador,lista_sku_no_metidos):
    
    # os.system('rm /Users/Fran/Desktop/cocinasplus/smeg/smeg.xlsx')
    # os.system('mv smeg/smeg_copia.xlsx /Users/Fran/Desktop/cocinasplus/smeg/smeg.xlsx')
    # os.system('cp smeg/smeg.xlsx /Users/Fran/Desktop/cocinasplus/smeg/smeg_copia.xlsx')

        
    
    excel_smeg='SMEG/smeg.xlsx'
   
    

        
    nombre_marca='SMEG'
        
    try:
        os.mkdir(nombre_marca+"/fotos")
        time.sleep(0.1)
    except:
        pass   
    
    contador_no_metidos=0 
    lista_sku_no_metidos=[]
    ean=''
    # l = input("pulse una tecla para continuar")
    wb = load_workbook(excel_smeg)
    ws = wb.active
    
    
    
    a_=ws['A2':'A'+str(ws.max_row)]#sku
    
    e_=ws['E2':'E'+str(ws.max_row)]#dimensiones
    g_=ws['G2':'G'+str(ws.max_row)]#ean
    h_=ws['H2':'H'+str(ws.max_row)]#precio_venta
    i_=ws['I2':'I'+str(ws.max_row)]#precio_costo
    n_=ws['N2':'N'+str(ws.max_row)]#categoria
    
    for i in range(0,int(ws.max_row)-1):
       
        try:
            time.sleep(0.1)
            
            sku_previo=a_[i][0].value
            sku='044.'+sku_previo
            nombre_foto=sku+'.jpg'
            categoria=n_[i][0].value
            
            categoria=categorias[str(categoria)]
            
            ean=str(g_[i][0].value)
            
            dimensiones=str(e_[i][0].value)
            dimensiones=re.sub("\(.*?\)","",dimensiones)
            dimensiones=re.sub("mm","",dimensiones)
            dimensiones=dimensiones.strip()
            alto=dimensiones.split("x")[0]
            largo=dimensiones.split("x")[1]
            ancho=dimensiones.split("x")[2]
            
            precio_coste=str(i_[i][0].value)
            precio_coste=precio_coste.replace('€','')
            precio_coste=precio_coste.strip()
            precio_coste=precio_coste.replace(',','.')
            precio_coste=float(precio_coste)*1.21
            precio_coste=round(precio_coste,2)
            
            precio_venta=str(h_[i][0].value)
            precio_venta=precio_venta.replace('€','')
            precio_venta=precio_venta.strip()
            precio_venta=precio_venta.replace(',','.')
            precio_venta=float(precio_venta)*1.21
            precio_venta=round(precio_venta,0)
            
            
            
            
            print(f'sku: {sku}')
            print(f'ean: {ean}')
            print(f'categoria: {categoria}')
            print(f'alto: {alto}, largo: {largo}, ancho: {ancho}')
            print(f'precio_coste: {precio_coste}')
            print(f'precio_venta: {precio_venta}')
            categoria=categoria.split('>')[-1]
            for i in categorias_: 
                categ=categorias_[i].split('>')[-1]
                categoria=categoria.split('>')[-1]
                if categ==categoria or categoria in categ:
                    categoria=i
                    print(categoria)
                    
                    break
                    
                    
            configurable=False
            
            try:
                
                try:
                    os.system('rm /Users/Fran/Desktop/cocinasplus/smeg/smeg.xlsx')
                    os.system('mv smeg/smeg_copia.xlsx /Users/Fran/Desktop/cocinasplus/smeg/smeg.xlsx')
                    os.system('cp smeg/smeg.xlsx /Users/Fran/Desktop/cocinasplus/smeg/smeg_copia.xlsx')
                except:
                    print("problema con el archivo smeg")
                    
                print("ahora la web")
                
                nombre_producto,descripcion, descripcion_corta = get_from_smeg_web(sku)
                
                
                try:
                    nombre_producto=corregir_texto(nombre_producto)
                except:
                    pass
                
                try:
                    nombre_producto_en=translator.translate(nombre_producto, dest='en').text
                    print(nombre_producto_en)
                except:
                    nombre_producto_en=nombre_producto
                    print(nombre_producto_en)
                try:
                    descripcion_en=translator.translate(descripcion, dest='en').text
                    print(descripcion_en)
                except:
                    descripcion_en=descripcion
                    print(descripcion_en)
                try:
                    descripcion_corta_en=translator.translate(descripcion_corta, dest='en').text
                    print(descripcion_corta_en)
                except:
                    descripcion_corta_en=descripcion_corta
                    print(descripcion_corta_en)
                
                    
            except:
                lista_sku_no_metidos.append(sku)
                time.sleep(1)    
            
            
            peso=0   
            try:
                insert_campos(nombre_marca,sku,ean,nombre_marca,nombre_producto,nombre_foto,largo, ancho,alto, peso,precio_coste,precio_venta,descripcion,descripcion,0,categoria)
                insert_campos_en(nombre_marca,sku,ean,nombre_marca,nombre_producto_en,nombre_foto,largo, ancho,alto, peso,precio_coste,precio_venta,descripcion,descripcion_en,0,categoria)
            except:
                pass    
            print('\n')
            print(f'..producto {sku} insertado correctamente en nuestra base de datos')
            now = datetime.now()
            now=now.strftime("%Y-%m")
            year=now[:4]
            mes=now[5:]
            print('9')
            nombre_marca='smeg'
            
            try:
            
                try:
                    nombre_marca_parafoto=nombre_marca.split(' ')[0]
                except:
                    pass
                
                path='/Users/Fran/Desktop/cocinasplus/smeg/fotos/'+nombre_foto
                try:
                    os.system('bash send_foto.sh '+path+' '+nombre_foto+' '+mes +' '+year)
                except:
                    path='/Users/Fran/Desktop/cocinasplus/smeg/fotos/'+nombre_foto.lower()
                    try:
                        os.system('bash send_foto.sh '+path+' '+nombre_foto+' '+mes +' '+year)
                    except:
                        pass
                foto='https://www.pluscookingdeluxe.com/wp-content/uploads/'+str(year)+'/'+str(mes)+'/'+str(nombre_foto)
                print(f'..foto {nombre_foto} subida correctamente a nuestro servidor')
                os.system('bash cambiar_permisos.sh www.pluscookingdeluxe.com/wp-content/uploads/'+str(year)+'/'+str(mes)+'/'+str(nombre_foto))
                print("permisos cambiados")
                time.sleep(0.5)
                
                print('10')
                
                slug=nombre_marca+'-'+sku+'-'+nombre_producto
                
                permalink="https://www.pluscookingdeluxe.com/product/"+nombre_marca+'-'+sku+'-'+nombre_producto+"/"
                nombre_producto=nombre_producto.split('(')[0]
                print(nombre_foto)
                data = {"name": nombre_producto,
                        "slug": slug,
                        "permalink": permalink,
                        "type": "simple",
                        "regular_price": str(precio_venta),
                        "cost": str(precio_coste),
                        "description": descripcion,
                        "short_description": descripcion_corta,
                        "sku": sku,
                        "ean": ean,
                
                        "stock_quantity": 1,
                        "manage_stock": True,
                        "backorders": "notify",
                        "weight": str(peso),
                        "dimensions":{
                                        "length": str(largo),
                                        "width": str(ancho),
                                        "height": str(alto)
                                    },
                        "lang": "es",
                        "seo": {
                                "title": nombre_producto,
                                "description": descripcion_corta,
                                "focus_keyword": nombre_producto
                                },
                    
                        "categories": 
                            [{"id": categoria}],
                            
                        "images": [
                            {
                                "src": foto,
                                "name": nombre_foto,
                                "alt": nombre_producto
                            }
                        ],
                        "meta_data": [
                                {
                                "key": "_alg_wc_cog_cost",
                                "value": str(precio_coste)
                                }
                        ]
                        }
                print(data)
                print('11')
            
                try:
                    resumen=wcapi.post("products", data).json()
                    print(resumen)
                    if resumen['message']=='SKU no válido o duplicado.':
                        resumen=wcapi.put("products/"+str(resumen['data']['resource_id']), data).json()
                        print(resumen)
                except Exception as e:
                    print(e)
                    
                    
                
                print('\n')
                print(f'el nombre de la marca es {nombre_marca}')
                
                print('12')
                try:
                    marca=wcapi.put("products/"+str(resumen['id']), {"brands": 315}).json()
                except:
                    pass
                
                
                nombre_producto_en=nombre_producto_en.split('(')[0]
                print(nombre_producto_en)
                try:
                    data = {
                            'name': nombre_producto_en,
                            "slug": nombre_producto_en+'-'+sku+'-'+nombre_marca,
                            "permalink": "https://www.pluscookingdeluxe.com/product/"+nombre_producto_en+'-'+sku+'-'+nombre_marca+"/",
                            'type' : 'simple',
                            'description' : descripcion_en,
                            'lang' : 'en',
                            'short_description':descripcion_en[0:40]+'...',
                            'translation_of':str(resumen['id'])
                    }
                    op=0
                    while op < 20:
                        print('\n')
                        op +=1
                    resumen_en=wcapi.post("products", data).json()
                    print(resumen_en)
                except:
                    pass
                try:
                    marca=wcapi.put("products/"+str(resumen_en['id']), {"brands": 315}).json()
                except:
                    pass
                
                    time.sleep(5)

                
            except:
                print("errores han habido")
                contador_no_metidos=contador_no_metidos+1
                lista_sku_no_metidos.append(sku)
                continue
        except:
            print("errores han habido")
            contador_no_metidos=contador_no_metidos+1
            lista_sku_no_metidos.append(sku)
            continue
        print(f'contador no metidos: {contador_no_metidos}')
                
    print('\n')
    print(f'..marca {nombre_marca} insertada correctamente')
    wb.save(excel)
    print(f'contador no metidos: {contador_no_metidos}')
    print(f'lista sku no metidos: {lista_sku_no_metidos}')